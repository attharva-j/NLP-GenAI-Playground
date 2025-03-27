# from base import ParsingAdapter
# from configs import ParsingConfig
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
from typing import List, Dict, Any, Optional
import xlrd
import openpyxl


class ExcelAdapter():
    """Adapter for parsing Excel files into a plain text format."""

    def __init__(self, config: Optional[Dict] = None):
        """
        Initialize the parsing adapter.

        Args:
            config (Optional[ParsingConfig]): Configuration for the parser.
        """
        self.config = config or {}
        print("ExcelAdapter instance created.")
    def _connect(self):
        """Dummy method to satisfy abstract class requirements."""
        pass

    def unmerge_cells(self, sheet):
        """Unmerge cells and fill them with the original top-left value."""
        merged_ranges = list(sheet.merged_cells.ranges)  # Get all merged cell ranges
    
        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_value = sheet.cell(row=min_row, column=min_col).value  # Get top-left value
    
            # Unmerge the cells before assigning values
            sheet.unmerge_cells(str(merged_range))
    
            # Assign the top-left value to all previously merged cells
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row=row, column=col).value = top_left_value

    def extract_data(self, wb, sheet_name):
        """Extract structured data from a given sheet."""
        sheet = wb[sheet_name]
        self.unmerge_cells(sheet)  # Normalize merged cells
        
        # Convert sheet to dataframe
        df = pd.DataFrame(sheet.values)
        
        # Drop empty columns and rows
        df.dropna(how='all', axis=0, inplace=True)
        df.dropna(how='all', axis=1, inplace=True)
        
        # Identify potential text sections (headers, footnotes)
        text_sections = []
        for row in df.itertuples(index=False):  # Avoid including row index
            # print(f"Row in df: {row}")
            row_values = [str(cell).strip() for cell in row if pd.notna(cell) and cell is not None]
            # print(f"Row values in df: {row_values}")
            if len(row_values) == 1:  # Possible title or note
                text_sections.append(row_values[0])
    
        # Convert table into structured format, removing `None` values
        table_data = df.fillna("").astype(str).to_dict(orient="records")  # Fill NaNs with empty strings

        
        return text_sections, table_data
        
    
    def _convert_xls_to_xlsx(self, binary_data: bytes) -> bytes:
        """Converts .xls to .xlsx in-memory."""
        input_stream = BytesIO(binary_data)
        
        # Load .xls file
        workbook_xls = xlrd.open_workbook(file_contents=input_stream.read())
        sheet_xls = workbook_xls.sheet_by_index(0)
    
        # Create new .xlsx workbook
        workbook_xlsx = openpyxl.Workbook()
        sheet_xlsx = workbook_xlsx.active
        sheet_xlsx.title = sheet_xls.name
    
        # Copy data from .xls to .xlsx
        for row in range(sheet_xls.nrows):
            for col in range(sheet_xls.ncols):
                sheet_xlsx.cell(row=row+1, column=col+1, value=sheet_xls.cell_value(row, col))
    
        # Save as .xlsx in-memory
        output_stream = BytesIO()
        workbook_xlsx.save(output_stream)
        
        return output_stream.getvalue()

    def parse(self, data: bytes, metadata: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
        """Parse the Excel file and return a list of dictionaries, each representing a sheet.

        Args:
            data (bytes): The binary data to be parsed.
            metadata (Optional[Dict[str, Any]]): Additional metadata about the file.

        Returns:
            List[Dict[str, Any]]: List of parsed content and metadata per sheet.
        """
        try:
            print(f"Parsing file with metadata {str(metadata)}")

            file_name = metadata.get("file_name", "input.xls")
            file_format = file_name.lower().split(".")[-1]

            if file_format == "xls":
                data = self._convert_xls_to_xlsx(data)

            wb = load_workbook(filename=BytesIO(data), data_only=True)

            result = []

            for sheet_name in wb.sheetnames:
                text_sections, table_data = self.extract_data(wb, sheet_name)

                sheet_text = []

                # Add text sections
                sheet_text.extend(text_sections)

                # Add table data as text
                for record in table_data:
                    text_entry = " ".join(f"{value}" for key, value in record.items())
                    sheet_text.append(text_entry)

                # Create one entry per sheet
                sheet_result = {
                    "content": "\n".join(sheet_text),
                    "metadata": {
                        **(metadata or {}),
                        "sheet_name": sheet_name,
                        "file_type": 'excel_sheet'
                    }
                }

                result.append(sheet_result)

            return result

        except Exception as e:
            return [{"error": str(e), "metadata": metadata or {}}]


        